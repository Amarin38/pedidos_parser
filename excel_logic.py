import pandas as pd
import regex as re
import zipfile
from io import BytesIO
from typing import Any, List, Tuple

import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Side, Border

import constants as const
from constants import TipoPedidoEnum, SepararPorEnum, FormatoPedidoEnum
from db.repositories.codigos_repository import CodigosRepository
from db.vm.codigos_vm import CodigosVM


class Pedidos:
    def __init__(self) -> None:
        self.repo = CodigosRepository()
        self.vm = CodigosVM()
        self.cod = Codigos()

    def ejecutar_todo(self, archivos_raw: List[Any], separar_por: SepararPorEnum):
        lista_df = self.limpiar(archivos_raw)
        todos_los_excels = []

        match separar_por:
            case SepararPorEnum.TODO:
                todos_los_excels.extend(self.filtrar(lista_df, archivos_raw, TipoPedidoEnum.OFICINA))
                todos_los_excels.extend(self.filtrar(lista_df, archivos_raw, TipoPedidoEnum.FLAVIO))

            case SepararPorEnum.SOLO_OFICINA:
                todos_los_excels.extend(self.filtrar(lista_df, archivos_raw, TipoPedidoEnum.OFICINA))

            case SepararPorEnum.SOLO_FLAVIO:
                todos_los_excels.extend(self.filtrar(lista_df, archivos_raw, TipoPedidoEnum.FLAVIO))

            case SepararPorEnum.SOLO_ROPA:
                todos_los_excels.extend(self.filtrar(lista_df, archivos_raw, TipoPedidoEnum.ROPA))

        if separar_por != SepararPorEnum.SOLO_ROPA:
            todos_los_excels.extend(self.filtrar_resto(lista_df, archivos_raw))

        return self.crear_zip(todos_los_excels)
    

    def limpiar(self, archivos_raw: List[Any]) -> List[Tuple[pd.DataFrame,str]]:
        lista_df = []

        for archivo_raw in archivos_raw:
            try:
                df = pd.read_fwf(
                        archivo_raw,
                        encoding=const.WS_ENCODING, 
                        skiprows=5
                    )
                
                nombre_archivo: str = archivo_raw.name
                del_desde = len(df) - 1
                del_hasta = del_desde - 1

                df = df.drop([0, del_desde, del_hasta]) # elimino todo, dejo solo la tabla principal
                df = df.rename(columns=const.COLUMNAS)

                df[const.COL_ARTICULO]      = df[const.COL_COLUMNAS].str.extract(const.REGEX_ARTICULO) # extraigo los articulos
                df[const.COL_COD_PROVEEDOR] = df[const.COL_COLUMNAS].str.extract(const.REGEX_COD_PROV) # extraigo los codigos proveedor
                
                df[const.COL_DESCRIPCION] = df.apply(self.extraer_dinamico, 
                                             axis=1, 
                                             args=(df, const.COL_COD_PROVEEDOR, 12, 40)) #extraigo la descripcion
                
                df[const.COL_CANTIDAD] = df.apply(self.extraer_dinamico, 
                                          axis=1, 
                                          args=(df, const.COL_DESCRIPCION, 40, 85)) # extraigo la cantidad

                df[const.COL_CANTIDAD] = df[const.COL_CANTIDAD].astype(str).str.extract(const.REGEX_CANTIDAD, expand=False)
                df[const.COL_CANTIDAD] = pd.to_numeric(df[const.COL_CANTIDAD], errors='coerce')

                df = df.drop(columns=const.COL_COLUMNAS)
                lista_df.append((df, nombre_archivo))
            except Exception as e:
                print(f"No se puede procesar el archivo {archivo_raw.name} -> {e}")
        return lista_df


    def filtrar(self, lista_df: List[Tuple[pd.DataFrame,str]], archivos_raw: List[Any], tipo_filtro: TipoPedidoEnum) -> List[Tuple[pd.DataFrame,str]]:
        excels_generados = []

        try:
            codigos = self.cod.separar_codigos(tipo_filtro)[const.COL_CODIGOS]

            for (df, nombre_archivo), archivo_raw in zip(lista_df, archivos_raw):
                df[const.COL_ARTICULO] = df[const.COL_ARTICULO].astype(str).str.strip()
                pertenece = df[const.COL_ARTICULO].isin(codigos)
                df_final = df[pertenece]
                
                nombre, formato = self.regex_pedido(nombre_archivo)
                nombre_final = f"PEDIDO {formato} {tipo_filtro} {nombre}.xlsx"
                wb = self.formatear_excel(df_final, archivo_raw, formato, tipo_filtro)

                if wb is not None:
                    excels_generados.append((wb, nombre_final))
        except FileNotFoundError as e:
            print("No existe el archivo -> ", e)

        return excels_generados
        

    def filtrar_resto(self, lista_df: List[Tuple[pd.DataFrame,str]], archivos_raw: List[Any]) -> List[Tuple[pd.DataFrame,str]]:
        codigos_flavio  = self.cod.separar_codigos(TipoPedidoEnum.FLAVIO)[const.COL_CODIGOS]
        codigos_oficina = self.cod.separar_codigos(TipoPedidoEnum.OFICINA)[const.COL_CODIGOS]
        codigos_ropa    = self.cod.separar_codigos(TipoPedidoEnum.ROPA)[const.COL_CODIGOS]
        
        lista_resto = []

        for (df, nombre_archivo), archivo_raw in zip(lista_df, archivos_raw):
            todos           = pd.concat([codigos_flavio, codigos_oficina, codigos_ropa]).astype(str).str.strip()
            no_pertenece    = ~df[const.COL_ARTICULO].isin(todos)
            df_resto        = df[no_pertenece]

            nombre, formato = self.regex_pedido(nombre_archivo)
            nombre_salida = f"RESTO PEDIDO {formato} {nombre}.xlsx"
            wb = self.formatear_excel(df_resto, archivo_raw, formato, TipoPedidoEnum.RESTO)

            tupla_resto = (wb, nombre_salida)

            if tupla_resto and tupla_resto[0] is not None:
                lista_resto.append(tupla_resto)
        return lista_resto


    def formatear_excel(self, df: pd.DataFrame, archivo_raw, formato_pedido: FormatoPedidoEnum, pedido_para: TipoPedidoEnum | None = None) -> openpyxl.Workbook | None:
        if df is None or df.empty:
            return None
        
        thick_side = Side(border_style=const.WS_BORDER_THICK, color=const.WS_BORDER_COLOR)
        thin_side = Side(border_style=const.WS_BORDER_THIN, color=const.WS_BORDER_COLOR)

        border_style_datos = Border(top=thick_side, left=thick_side, right=thick_side, bottom=thick_side)
        border_style_tabla = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)
        
        archivo_raw.seek(0)
        primera_linea = archivo_raw.readline().decode(const.WS_ENCODING).strip()

        lineas_encabezado = []

        for _ in range(4):
            linea_bytes = archivo_raw.readline()
            lineas_encabezado.append(linea_bytes.decode(const.WS_ENCODING))

        texto_encabezado = "".join(lineas_encabezado)
        datos = self.filtrar_datos(texto_encabezado, formato_pedido)

        workbook = openpyxl.Workbook() # cargo el excel
        work_sheet = workbook.active # selecciono la hoja activa

        if work_sheet is not None:
            work_sheet.title = const.WS_TITLE

            if formato_pedido == FormatoPedidoEnum.NORMAL:
                work_sheet['A1'] = const.WS_NOTA_PED  
                work_sheet['A4'] = const.WS_R_SOCIAL 
            elif formato_pedido == FormatoPedidoEnum.PENDIENTE:
                work_sheet['A1'] = const.WS_PED_PEN  
                work_sheet['A4'] = const.WS_PED_FECHA 
            
            work_sheet['A2'] = const.WS_PROVEEDOR 
            work_sheet['A3'] = const.WS_FECHA 
            work_sheet['A5'] = const.WS_PARA 
            work_sheet['A6'] = const.WS_PEDIDO_PARA
            work_sheet['A7'] = const.WS_TIPO_PEDIDO
            
            work_sheet['B1'] = datos[0] # num_pedido 
            work_sheet['B2'] = datos[3] # proveedor
            work_sheet['B3'] = datos[1] # fecha
            work_sheet['B4'] = datos[4] # razon_social
            work_sheet['B5'] = f"{primera_linea} - {datos[2]}" # para_quien
            work_sheet['B6'] = pedido_para
            work_sheet['B7'] = formato_pedido

            work_sheet.append([])

            for fila in dataframe_to_rows(df, index=False, header=True):
                work_sheet.append(fila)

            work_sheet.column_dimensions['A'].width = const.WS_A_WIDTH
            work_sheet.column_dimensions['B'].width = const.WS_B_WIDTH
            work_sheet.column_dimensions['C'].width = const.WS_C_WIDTH
            work_sheet.column_dimensions['D'].width = const.WS_D_WIDTH
            
            # DATOS
            for row in work_sheet.iter_rows(min_row=1, max_row=7, min_col=1, max_col=2):
                for cell in row:
                    cell.border = border_style_datos
                    
            for row in work_sheet.iter_rows(min_row=1, max_row=7, min_col=1, max_col=1):
                for cell in row:
                    cell.font = Font(bold=True)


            # TABLA
            for cell in work_sheet[9]: 
                cell.font = Font(bold=True)

            for row in work_sheet.iter_rows(min_row=9, max_row=9, min_col=1, max_col=4):
                for cell in row:
                    cell.border = border_style_datos
            
            for row in work_sheet.iter_rows(min_row=10, 
                                            max_row=work_sheet.max_row, 
                                            min_col=1, 
                                            max_col=work_sheet.max_column):
                for cell in row:
                    cell.border = border_style_tabla

            return workbook
        return openpyxl.Workbook()


    def crear_zip(self, lista_workbooks: List[openpyxl.Workbook]):
        zip_buffer = BytesIO()

        with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
            for item in lista_workbooks:
                if item is None:
                    continue
                
                workbook, nombre_archivo = item

                if isinstance(item, tuple) and len(item) == 2:
                    workbook, nombre_archivo = item

                    if workbook is None:
                        continue

                    excel_buffer = BytesIO()
                    workbook.save(excel_buffer)
                    zip_file.writestr(nombre_archivo, excel_buffer.getvalue())
        return zip_buffer.getvalue()


    @staticmethod
    def extraer_dinamico(fila, df, nombre_col_base, val_vacio, val_texto):
        texto_linea = str(fila[const.COL_COLUMNAS])
        valor_celda = str(fila[nombre_col_base]) if pd.notna(fila[nombre_col_base]) else ""
        largo_celda = len(valor_celda)
        
        if largo_celda == 0:
            texto_final = texto_linea[val_vacio:].strip()
        else:
            texto_final = texto_linea[val_texto:].strip()

        match = re.search(const.REGEX_EXTRAER_DIN, texto_final)

        if match:
            return match.group(0).strip()
        else:
            return texto_final


    @staticmethod
    def regex_pedido(nombre_archivo: str):
        es_pendiente = re.search(const.PED_PEN, nombre_archivo)
        match_nombre = re.search(const.REGEX_PED_PEN, nombre_archivo) if es_pendiente else re.search(const.REGEX_PED, nombre_archivo)
        formato      = FormatoPedidoEnum.PENDIENTE if es_pendiente else FormatoPedidoEnum.NORMAL
        
        return (match_nombre.group(1) if match_nombre is not None else None, formato)


    @staticmethod
    def filtrar_datos(texto_encabezado: str, formato_pedido: FormatoPedidoEnum) -> tuple[str | Any, ...]:
        # 1. NOTA DE PEDIDO (Busca dígitos continuos después de los dos puntos)
        if formato_pedido == FormatoPedidoEnum.NORMAL:
            match_pedido = re.search(const.REGEX_NOTA_PED, texto_encabezado)
        elif formato_pedido == FormatoPedidoEnum.PENDIENTE:
            match_pedido = re.search(const.REGEX_NOTA_PED_PEN, texto_encabezado)
            
        num_pedido = match_pedido.group(1) if match_pedido else const.NO_DATA

        # 2. FECHA (Busca números y barras)
        match_fecha = re.search(const.REGEX_FECHA, texto_encabezado)
        fecha = match_fecha.group(1) if match_fecha else const.NO_DATA

        # 3. Para (Captura todo hasta encontrar 2+ espacios seguidos o salto de línea)
        match_para = re.search(const.REGEX_PARA, texto_encabezado)
        para_quien = match_para.group(1).strip() if match_para else const.NO_DATA

        # 4. PROVEEDOR (Igual que el anterior)
        match_proveedor = re.search(const.REGEX_PROVEEDOR, texto_encabezado)
        proveedor = match_proveedor.group(1).strip() if match_proveedor else const.NO_DATA

        # 5. R. SOCIAL 
        # (Nota: hay que escapar el punto de R\. SOCIAL con una barra invertida)
        if formato_pedido == FormatoPedidoEnum.NORMAL:
            match_rsocial = re.search(const.REGEX_RAZON_SOCIAL, texto_encabezado)
        elif formato_pedido == FormatoPedidoEnum.PENDIENTE:
            match_rsocial = re.search(const.REGEX_PED_FECHAS, texto_encabezado)

        razon_social = match_rsocial.group(1).strip() if match_rsocial else const.NO_DATA

        return num_pedido, fecha, para_quien, proveedor, razon_social 



class Codigos:
    def __init__(self) -> None:
        self.repo = CodigosRepository()
        self.vm = CodigosVM()


    def sacar_lista(self) -> None:
        df = pd.read_excel(const.CODIGOS_PATH, dtype={const.COL_CODIGO: str})
        
        df[const.COL_CODIGO]    = df[const.COL_CODIGO].astype(str)
        df[const.COL_ARTICULO]  = df[const.COL_ARTICULO].astype(str)

        df = df.rename(columns={const.COL_ARTICULO : const.COL_DESCRIPCION})
        
        df[[const.COL_FAMILIA, const.COL_ARTICULO]] = df[const.COL_CODIGO].str.strip().str.split(const.SEPARATOR, expand=True)
        
        df[const.COL_FAMILIA]   = df[const.COL_FAMILIA].astype(str)
        df[const.COL_ARTICULO]  = df[const.COL_ARTICULO].astype(str)

        df = df.drop(columns=const.COL_CODIGO)
        
        familia = df[const.COL_FAMILIA].str.zfill(3)
        articulo = df[const.COL_ARTICULO].str.zfill(5) 
        df[const.COL_CODIGOS] = familia.str.cat(articulo, sep=const.SEPARATOR)

        self.repo.insert_many(self.vm.to_model(df))


    def separar_codigos(self, tipo_pedido: TipoPedidoEnum) -> pd.DataFrame:
        df = self.repo.get_df()

        df[const.COL_FAMILIA]   = df[const.COL_FAMILIA].astype(pd.Int32Dtype())
        df[const.COL_ARTICULO]  = df[const.COL_ARTICULO].astype(pd.Int32Dtype())
        
        if tipo_pedido != TipoPedidoEnum.ROPA:
            df_final = df.loc[df[const.COL_DEPOSITO] == tipo_pedido, :]
        else:
            df_final = df.loc[df[const.COL_FAMILIA] == const.ROPA_FAM, :]

        df_final = df_final.sort_values(by=[const.COL_FAMILIA, const.COL_ARTICULO], ascending=[True, True])
        df_final = df_final.reset_index(drop=True)
        df_final = df_final.loc[:, ~df_final.columns.str.contains(const.UNNAMED)]
        
        fam_final                   = df_final[const.COL_FAMILIA].astype(str).str.zfill(3)
        art_final                   = df_final[const.COL_ARTICULO].astype(str).str.zfill(5)
        df_final[const.COL_CODIGOS] = fam_final.str.cat(art_final, sep=const.SEPARATOR)

        return df_final


