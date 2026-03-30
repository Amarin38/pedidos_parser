import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Side, Border

import pandas as pd
import regex as re
from pathlib import Path
from contants import CODIGOS_PATH, OFICINA_PATH, OFICINA_XLSX_PATH, CLI, VALUE_ERR_CLI
from contants import TipoPedido, FormatoPedido
from contants import COLUMNAS
from typing import Any

# TODO: queda que funcione con los pedidos pendientes
class Pedidos:
    def __init__(self) -> None:
        self.archivos_raw = sorted(list(OFICINA_PATH.rglob("Ped *.txt")))


    def limpiar(self) -> list[pd.DataFrame]: # type: ignore
        lista_df = []

        for archivo_raw in self.archivos_raw:
            try:
                df = pd.read_fwf(
                        archivo_raw, 
                        encoding='latin1', 
                        skiprows=5
                    )
                
                file_size = len(df)

                df = df.drop([0, file_size-1, file_size-2]) # elimino todo, dejo solo la tabla principal
                df = df.rename(columns=COLUMNAS)

                df["Articulo"] = df["Columnas"].str.extract(r'^.{0}(.*?) \s{2,}') # extraigo los articulos
                df["Codigo Proveedor"] = df["Columnas"].str.extract(r'^.{12}(.*?) \s{2,}') # extraigo los codigos proveedor
                
                df["Descripcion"] = df.apply(self.extraer_dinamico, 
                                             axis=1, 
                                             args=(df, "Codigo Proveedor", 12, 40)) #extraigo la descripcion
                
                df["Cantidad"] = df.apply(self.extraer_dinamico, 
                                          axis=1, 
                                          args=(df, "Descripcion", 40, 85)) # extraigo la cantidad

                df["Cantidad"] = df["Cantidad"].astype(str).str.extract(r'(\d+\.?\d*)', expand=False)
                df["Cantidad"] = pd.to_numeric(df["Cantidad"], errors='coerce')

                df = df.drop(columns="Columnas")


                if re.search(r'Ped Pen', archivo_raw.name):
                    nombre = re.search(r'(Ped Pen.*?\d{2}-\d{2}-\d{4})', archivo_raw.name)

                    nombre_salida = OFICINA_XLSX_PATH / f"TODOS PENDIENTE {nombre.group(1) if nombre is not None else None}.xlsx"
                    print(f"Convertido: {archivo_raw.name} -> {nombre_salida.name}")

                    self.guardar_excel_formateado(df, archivo_raw, nombre_salida, FormatoPedido.PENDIENTE) 
                else:
                    nombre = re.search(r'(Ped.*?\d{2}-\d{2}-\d{4})', archivo_raw.name)

                    nombre_salida = OFICINA_XLSX_PATH / f"TODOS {nombre.group(1) if nombre is not None else None}.xlsx"
                    print(f"Convertido: {archivo_raw.name} -> {nombre_salida.name}")


                    self.guardar_excel_formateado(df, archivo_raw, nombre_salida, FormatoPedido.NORMAL) 
                
                lista_df.append(df)

            except Exception as e:
                print(f"No se puede procesar el archivo {archivo_raw.name} -> {e}")
        return lista_df


    def filtrar(self, lista_df: list[pd.DataFrame], tipo_filtro: TipoPedido) -> None:
        archivos_formateados = sorted(list(OFICINA_XLSX_PATH.rglob("TODOS *.xlsx")))
        
        try:
            codigos = pd.read_excel(CODIGOS_PATH / f"CODIGOS_{tipo_filtro}.xlsx", dtype={"Codigos": str})["Codigos"]
            
            for df, archivo_form, archivo_raw in zip(lista_df, archivos_formateados, self.archivos_raw):
                df["Articulo"] = df["Articulo"].astype(str).str.strip()
                pertenece = df["Articulo"].isin(codigos)
                df_final = df[pertenece]


                path_salida = Path()

                if re.search(r'Ped Pen', archivo_form.name):
                    nombre = re.search(r'(Ped Pen.*?\d{2}-\d{2}-\d{4})', archivo_form.name)

                    path_salida = OFICINA_XLSX_PATH / f"PEDIDO PENDIENTE {tipo_filtro} {nombre.group(1) if nombre is not None else None}.xlsx"

                    self.filtrar_resto(df, nombre, archivo_raw, FormatoPedido.PENDIENTE)
                    self.guardar_excel_formateado(df_final, archivo_raw, path_salida, FormatoPedido.PENDIENTE, tipo_filtro)
                else:
                    nombre = re.search(r'(Ped.*?\d{2}-\d{2}-\d{4})', archivo_form.name)

                    path_salida = OFICINA_XLSX_PATH / f"PEDIDO {tipo_filtro} {nombre.group(1) if nombre is not None else None}.xlsx"

                    self.filtrar_resto(df, nombre, archivo_raw, FormatoPedido.NORMAL)
                    self.guardar_excel_formateado(df_final, archivo_raw, path_salida, FormatoPedido.NORMAL, tipo_filtro)

        except FileNotFoundError as e:
            print("No existe el archivo -> ", e)


    def filtrar_resto(self, df: pd.DataFrame, nombre_archivo, archivo_raw: Path, formato_pedido: FormatoPedido) -> None:
        codigos_flavio  = pd.read_excel(CODIGOS_PATH / f"CODIGOS_{TipoPedido.FLAVIO}.xlsx", dtype={"Codigos": str})["Codigos"]
        codigos_oficina = pd.read_excel(CODIGOS_PATH / f"CODIGOS_{TipoPedido.OFICINA}.xlsx", dtype={"Codigos": str})["Codigos"]
        codigos_ropa    = pd.read_excel(CODIGOS_PATH / f"CODIGOS_{TipoPedido.ROPA}.xlsx", dtype={"Codigos": str})["Codigos"]
        
        todos = pd.concat([codigos_flavio, codigos_oficina, codigos_ropa])
        no_pertenece = ~df["Articulo"].isin(todos)

        df_resto = df[no_pertenece]
        
        nombre_salida = OFICINA_XLSX_PATH / f"RESTO PEDIDO {formato_pedido} {nombre_archivo.group(1)}.xlsx"

        self.guardar_excel_formateado(df_resto, archivo_raw, nombre_salida, formato_pedido, TipoPedido.RESTO)


    @staticmethod
    def extraer_dinamico(fila, df, nombre_col_base, val_vacio, val_texto):
        texto_linea = str(fila["Columnas"])
        valor_celda = str(fila[nombre_col_base]) if pd.notna(fila[nombre_col_base]) else ""
        largo_celda = len(valor_celda)
        
        if largo_celda == 0:
            texto_final = texto_linea[val_vacio:].strip()
        else:
            texto_final = texto_linea[val_texto:].strip()

        match = re.search(r'^.*?(?=\s{2,})', texto_final)

        if match:
            return match.group(0).strip()
        else:
            return texto_final


    @staticmethod
    def filtrar_datos(texto_encabezado: str, formato_pedido: FormatoPedido) -> tuple[str | Any, ...]:
        # 1. NOTA DE PEDIDO (Busca dígitos continuos después de los dos puntos)
        if formato_pedido == FormatoPedido.NORMAL:
            match_pedido = re.search(r'NOTA DE PEDIDO\s*:\s*(\d+)', texto_encabezado)
        elif formato_pedido == FormatoPedido.PENDIENTE:
            match_pedido = re.search(r'PEDIDO PENDIENTE\s*:\s*(\d+)', texto_encabezado)
            
        num_pedido = match_pedido.group(1) if match_pedido else "Sin datos"

        # 2. FECHA (Busca números y barras)
        match_fecha = re.search(r'FECHA\s*:\s*([\d/]+)', texto_encabezado)
        fecha = match_fecha.group(1) if match_fecha else "Sin datos"

        # 3. Para (Captura todo hasta encontrar 2+ espacios seguidos o salto de línea)
        match_para = re.search(r'Para:\s*(.*?)(?:\s{2,}|\n|$)', texto_encabezado)
        para_quien = match_para.group(1).strip() if match_para else "Sin datos"

        # 4. PROVEEDOR (Igual que el anterior)
        match_proveedor = re.search(r'PROVEEDOR\s*:\s*(.*?)(?:\s{2,}|\n|$)', texto_encabezado)
        proveedor = match_proveedor.group(1).strip() if match_proveedor else "Sin datos"

        # 5. R. SOCIAL 
        # (Nota: hay que escapar el punto de R\. SOCIAL con una barra invertida)
        if formato_pedido == FormatoPedido.NORMAL:
            match_rsocial = re.search(r'R\.\s*SOCIAL\s*:\s*(.*?)(?:\s{2,}|\n|$)', texto_encabezado)
        elif formato_pedido == FormatoPedido.PENDIENTE:
            match_rsocial = re.search(r'Pedido de Fecha\s*(.*?)(?:\s{2,}|\n|$)', texto_encabezado)

        razon_social = match_rsocial.group(1).strip() if match_rsocial else "Sin datos"

        return num_pedido, fecha, para_quien, proveedor, razon_social 


    def guardar_excel_formateado(self, df: pd.DataFrame, archivo_raw: Path, ruta_salida: Path, formato_pedido: FormatoPedido, pedido_para: TipoPedido | None = None) -> pd.DataFrame:
        if df.empty:
            return pd.DataFrame()
        
        thick_side = Side(border_style="thick", color="000000")
        thin_side = Side(border_style="thin", color="000000")
        border_style_datos = Border(top=thick_side, left=thick_side, right=thick_side, bottom=thick_side)
        border_style_tabla = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)
        
        with open(archivo_raw, 'r', encoding='latin1') as f:
            primera_linea = f.readline().strip()
            texto_encabezado = "".join([f.readline() for _ in range(4)])
            datos = self.filtrar_datos(texto_encabezado, formato_pedido)

        workbook = openpyxl.Workbook() # cargo el excel
        work_sheet = workbook.active # selecciono la hoja activa

        if work_sheet is not None:
            work_sheet.title = "PEDIDO" 

            if formato_pedido == FormatoPedido.NORMAL:
                work_sheet['A1'] = "NOTA DE PEDIDO:"  
                work_sheet['A4'] = "R.SOCIAL:" 
            elif formato_pedido == FormatoPedido.PENDIENTE:
                work_sheet['A1'] = "PEDIDO PENDIENTE:"  
                work_sheet['A4'] = "PEDIDO DE FECHA:" 
            
            work_sheet['A2'] = "PROVEEDOR:" 
            work_sheet['A3'] = "FECHA:" 
            work_sheet['A5'] = "PARA:" 
            work_sheet['A6'] = "PEDIDO PARA:"
            work_sheet['A7'] = "TIPO DE PEDIDO:"
            
            work_sheet['B1'] = datos[0] # num_pedido 
            work_sheet['B2'] = datos[3] # proveedor
            work_sheet['B3'] = datos[1] # fecha
            work_sheet['B4'] = datos[4] # razon_social
            work_sheet['B5'] = f"{primera_linea} - {datos[2]}" # para_quien
            work_sheet['B6'] = pedido_para
            work_sheet['B7'] = formato_pedido

            work_sheet.append([])

            for fila in dataframe_to_rows(df, index=False, header=True):
                work_sheet.append(fila)

            work_sheet.column_dimensions['A'].width = 25
            work_sheet.column_dimensions['B'].width = 45
            work_sheet.column_dimensions['C'].width = 50
            work_sheet.column_dimensions['D'].width = 15
            
            # DATOS
            for row in work_sheet.iter_rows(min_row=1, max_row=7, min_col=1, max_col=2):
                for cell in row:
                    cell.border = border_style_datos
                    
            for row in work_sheet.iter_rows(min_row=1, max_row=7, min_col=1, max_col=1):
                for cell in row:
                    cell.font = Font(bold=True)


            # TABLA
            for cell in work_sheet[9]: 
                cell.font = Font(bold=True)

            for row in work_sheet.iter_rows(min_row=9, max_row=9, min_col=1, max_col=4):
                for cell in row:
                    cell.border = border_style_datos
            
            for row in work_sheet.iter_rows(min_row=10, 
                                            max_row=work_sheet.max_row, 
                                            min_col=1, 
                                            max_col=work_sheet.max_column):
                for cell in row:
                    cell.border = border_style_tabla


            workbook.save(ruta_salida)
        return pd.DataFrame()


class Codigos:
    def __init__(self) -> None:
        pass


    def limpiar(self) -> None:
        df = pd.read_excel(CODIGOS_PATH / "codigos.xlsx", dtype={"Codigo": str})

        df["Codigo"] = df["Codigo"].astype(str)
        df["Articulo"] = df["Articulo"].astype(str)
        df = df.rename(columns={"Articulo":"Descripcion"})

        df[["Familia", "Articulo"]] = df["Codigo"].str.strip().str.split(".", expand=True)
        df["Familia"] = df["Familia"].astype(str)
        df["Articulo"] = df["Articulo"].astype(str)
        df = df.drop(columns="Codigo")

        familia = df["Familia"].str.zfill(3)
        articulo = df["Articulo"].str.zfill(5) 
        df["Codigos"] = familia.str.cat(articulo, sep=".")

        df.to_excel(CODIGOS_PATH / "CODIGOS_FINAL.xlsx")


    def sacar_lista(self) -> None:
        df = pd.read_excel(CODIGOS_PATH / "CODIGOS_FINAL.xlsx")

        df["Familia"] = df["Familia"].astype('Int32')
        df["Articulo"] = df["Articulo"].astype('Int32')


        df_flavio = df.loc[df["Deposito"] == "FLAVIO", :]
        df_flavio = df_flavio.sort_values(by=["Familia", "Articulo"], ascending=[True, True])
        df_flavio = df_flavio.reset_index(drop=True)
        df_flavio = df_flavio.loc[:, ~df_flavio.columns.str.contains("^Unnamed")]
        
        fam_flavio = df_flavio["Familia"].astype(str).str.zfill(3)
        art_flavio = df_flavio["Articulo"].astype(str).str.zfill(5) 
        df_flavio["Codigos"] = fam_flavio.str.cat(art_flavio, sep=".")

        df_oficina = df.loc[df["Deposito"] == "OFICINA", :]
        df_oficina = df_oficina.sort_values(by=["Familia", "Articulo"], ascending=[True, True])
        df_oficina = df_oficina.reset_index(drop=True)
        df_oficina = df_oficina.loc[:, ~df_oficina.columns.str.contains("^Unnamed")]
        
        fam_ofi = df_oficina["Familia"].astype(str).str.zfill(3)
        art_ofi = df_oficina["Articulo"].astype(str).str.zfill(5) 
        df_oficina["Codigos"] = fam_ofi.str.cat(art_ofi, sep=".")

        df_ropa = df.loc[df["Familia"] == 130, :]
        df_ropa = df_ropa.sort_values(by=["Familia", "Articulo"], ascending=[True, True])
        df_ropa = df_ropa.reset_index(drop=True)
        df_ropa = df_ropa.loc[:, ~df_ropa.columns.str.contains("^Unnamed")]

        fam_ropa = df_ropa["Familia"].astype(str).str.zfill(3)
        art_ropa = df_ropa["Articulo"].astype(str).str.zfill(5) 
        df_ropa["Codigos"] = fam_ropa.str.cat(art_ropa, sep=".")

        pd.DataFrame(df_flavio).to_excel(CODIGOS_PATH / f"CODIGOS_{TipoPedido.FLAVIO}.xlsx")
        pd.DataFrame(df_oficina).to_excel(CODIGOS_PATH / f"CODIGOS_{TipoPedido.OFICINA}.xlsx")
        pd.DataFrame(df_ropa).to_excel(CODIGOS_PATH / f"CODIGOS_{TipoPedido.ROPA}.xlsx")


if __name__ == '__main__':
    pedido = Pedidos()
    codigos = Codigos()

    codigos.limpiar()
    codigos.sacar_lista()

    print(CLI)
    tipo_filtrado = int(input(">> "))

    while tipo_filtrado != 0:
        try:
            match tipo_filtrado:
                case 0:
                    break
                case 1:
                    df = pedido.limpiar()

                    pedido.filtrar(df, TipoPedido.OFICINA)
                    pedido.filtrar(df, TipoPedido.FLAVIO)
                    pedido.filtrar(df, TipoPedido.ROPA)
                case 2: 
                    df = pedido.limpiar()

                    pedido.filtrar(df, TipoPedido.OFICINA)
                case 3: 
                    df = pedido.limpiar()

                    pedido.filtrar(df ,TipoPedido.FLAVIO)
                case 4:
                    df = pedido.limpiar()

                    pedido.filtrar(df, TipoPedido.ROPA)
        
            print(CLI)
            tipo_filtrado = int(input(">> "))    
        except ValueError as e:
            print(VALUE_ERR_CLI, e)

  

    

